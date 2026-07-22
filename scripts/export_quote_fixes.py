#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Export the quotations that restore_quran_quotes.py could not restore.

    python scripts/export_quote_fixes.py            write the workbook
    python scripts/export_quote_fixes.py --ingest   read the filled-in one

The workbook (raw/quote_fixes.xlsx) has one row per [surah:ayah] reference
that was left as the book printed it, with where to find it — the root article
and the printed page, located in the PDF — and blank columns for the correct
reference and, if needed, the correct Arabic.

--ingest turns the filled-in workbook into raw/quote_fixes.json, which is what
the build actually reads. The two files are kept apart deliberately: the
workbook is a working document that can be regenerated, the JSON is the record
of every correction and is never overwritten by an export.
"""
import collections
import io
import itertools
import json
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import build_root_dictionary
import restore_quran_quotes as R

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)
PDF = os.path.join(REPO, 'raw', 'Fatuhat-al-Quran- Final Draft.pdf')
OUT = os.path.join(REPO, 'raw', 'quote_fixes.xlsx')

REASON = {
    'out-of-range': 'Reference does not exist',
    'too-short': 'Quote too short to match safely',
    'ambiguous': 'Wording occurs in several ayaat',
    'partial': 'Only part of the quote is in the cited ayah',
    'no-overlap': 'Quote is not in the cited ayah at all',
}
ORDER = ['Reference does not exist', 'Wording occurs in several ayaat',
         'Quote is not in the cited ayah at all',
         'Only part of the quote is in the cited ayah',
         'Quote too short to match safely']


# --------------------------------------------------------------- the book

def index_pdf():
    """Printed page number, headwords and references, per PDF page."""
    import fitz
    import re

    letters = 'ء-غف-يىةکھہۃیٱ'
    spaced = re.compile('(?<![' + letters + '])([' + letters + ']'
                        '(?: [' + letters + ']){1,4})(?![' + letters + '])')
    ref = re.compile(r'(\d{1,3})\s*:\s*(\d{1,3})')
    marks = re.compile('[ً-ٰۖ-ۭـ]')
    fold = {'أ': 'ا', 'إ': 'ا', 'آ': 'ا', 'ٱ': 'ا', 'ى': 'ي', 'ی': 'ي',
            'ہ': 'ه', 'ھ': 'ه', 'ۃ': 'ه', 'ة': 'ه', 'ک': 'ك'}

    doc = fitz.open(PDF)
    pages = []
    for i in range(doc.page_count):
        raw = doc[i].get_text()
        txt = marks.sub('', raw)
        first = raw.strip().split('\n', 1)[0].strip()
        pages.append({
            'pdf': i + 1,
            'printed': int(first) if first.isdigit() else None,
            'heads': {''.join(fold.get(c, c) for c in m.group(1))
                      for m in spaced.finditer(txt)},
            'refs': {'%s:%s' % (m.group(1), m.group(2))
                     for m in ref.finditer(txt)},
        })
    # Pages that print no number inherit it from the page before.
    last = None
    for p in pages:
        if p['printed'] is not None:
            last = p['printed']
        elif last is not None:
            last += 1
            p['printed'] = last
    return pages


# --------------------------------------------------------------- matching

def run_in(ws, a_skel):
    """Longest suffix of the quote skeleton found anywhere in the ayah."""
    for length in range(min(len(ws), len(a_skel)), 0, -1):
        if ws[-length:] in a_skel:
            return length
    return 0


def suggest(ref, ws, skels, last):
    """The conversion scrambles the digits of a reference ([5:253] for 25:53).
    Try every rearrangement of the printed digits and keep the ayah that
    actually contains the most of the quote — the text is the evidence, the
    reference only narrows the field. Only returned when it beats every other
    rearrangement clearly, so a guess is never presented as a certainty."""
    digits = ''.join(c for c in ref if c.isdigit())
    if not digits or len(digits) > 6 or len(ws) < 4:
        return None, 0
    seen, best, second = set(), (0, None), 0
    for perm in set(itertools.permutations(digits)):
        d = ''.join(perm)
        for i in range(1, len(d)):
            s, a = d[:i], d[i:]
            if not s.lstrip('0') or not a.lstrip('0'):
                continue
            si, ai = int(s), int(a)
            if not (1 <= si <= 114 and 1 <= ai <= last[si]):
                continue
            aid = '%d:%d' % (si, ai)
            if aid in seen:
                continue
            seen.add(aid)
            n = run_in(ws, skels[aid])
            if n > best[0]:
                second, best = best[0], (n, aid)
            elif n > second:
                second = n
    if best[0] >= 5 and best[0] >= second + 2:
        return best[1], best[0]
    return None, 0


def collect():
    overrides = R.load_overrides()
    entries = build_root_dictionary.build_entries(verbose=False)[0]
    _out, st, log = R.restore_all(entries, overrides)
    ayaat, last = R.load_quran()
    skels = {aid: R.skel(t)[0] for aid, t in ayaat.items()}

    pages = index_pdf()
    head_pages = collections.defaultdict(list)
    ref_pages = collections.defaultdict(list)
    for p in pages:
        for h in p['heads']:
            head_pages[h].append(p)
        for r in p['refs']:
            ref_pages[r].append(p)

    def locate(root, ref):
        """The page carrying this reference inside the root's own article,
        falling back to where that article starts."""
        rp = head_pages.get(root, [])
        cand = ref_pages.get(ref.strip('[] ').replace(' ', ''), [])
        if rp:
            lo, hi = rp[0]['pdf'], rp[-1]['pdf'] + 2
            inside = [p for p in cand if lo <= p['pdf'] <= hi]
            if inside:
                return inside[0]['printed'], 'ref on page'
            return rp[0]['printed'], 'root article'
        if len(cand) == 1:
            return cand[0]['printed'], 'ref on page'
        return None, 'not found'

    rows = []
    for root, ref, _aid, kind, window, context in log:
        if kind != 'nomatch':
            continue
        sr, ar = ref.strip('[]').split(':')
        cands = R.reference_candidates(sr, ar, last)
        ws, _ = R.skel(window)

        best, where = 0, ''
        for s, a in cands:
            aid = '%d:%d' % (s, a)
            n = run_in(ws, skels[aid])
            if n > best:
                best, where = n, aid

        if not cands:
            why = 'out-of-range'
        elif len(ws) < R.MIN_CITED:
            why = 'too-short'
        else:
            hits = [aid for aid, sk in skels.items()
                    if len(ws) >= 12 and ws[-12:] in sk]
            if len(hits) > 1:
                why = 'ambiguous'
                where = ', '.join(hits[:6]) + ('…' if len(hits) > 6 else '')
            elif best >= 3:
                why = 'partial'
            else:
                why = 'no-overlap'

        page, how = locate(root, ref)
        sug, sug_n = suggest(ref, ws, skels, last)
        if sug and sug_n <= best:        # no better than the printed reference
            sug, sug_n = None, 0

        quote = ' '.join(window.split())
        # A correction that was made but did not line up with what the book
        # printed comes back with the row, so it can be looked at again.
        prev = overrides.get((root, ref, quote), ('', ''))

        rows.append({
            'why': REASON[why], 'root': root, 'page': page, 'how': how,
            'ref': ref.strip(), 'closest': where or '—', 'matched': best,
            'quote': quote,
            'context': ' '.join(context.split())[-160:],
            'suggest': sug or '', 'suggest_n': sug_n or '',
            'fix_ref': prev[0], 'fix_text': prev[1],
            'note': 'earlier correction did not line up with the printed quote'
                    if any(prev) else '',
        })

    rows.sort(key=lambda r: (ORDER.index(r['why']), r['page'] or 9999, r['root']))
    return rows, st


# --------------------------------------------------------------- workbook

COLS = [
    ('#', 6, None),
    ('Why it was left alone', 34, None),
    ('Root', 12, 'ar'),
    ('Book page', 10, None),
    ('Page found by', 13, None),
    ('Reference as printed', 18, None),
    ('Closest ayah we found', 22, None),
    ('Matched', 9, None),
    ('Quote as printed (corrupted)', 46, 'ar'),
    ('Urdu text just before it', 46, 'ur'),
    ('Our suggestion', 14, 'sug'),
    ('Suggestion match', 11, 'sug'),
    ('CORRECT reference (S:A)', 20, 'fill'),
    ('CORRECT Arabic text (optional)', 46, 'fill-ar'),
    ('Notes', 26, 'fill'),
]


def write(rows, st):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    head_fill = PatternFill('solid', fgColor='E8E2D4')
    fill_in = PatternFill('solid', fgColor='FFF6DC')
    fill_sug = PatternFill('solid', fgColor='E6F0E4')
    thin = Side(style='thin', color='D8D2C4')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Quotes to fix'
    ws.append([c[0] for c in COLS])
    for i, (_name, width, _kind) in enumerate(COLS, start=1):
        c = ws.cell(row=1, column=i)
        c.font = Font(bold=True, color='1F2430', size=11)
        c.fill = head_fill
        c.border = border
        c.alignment = Alignment(vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[1].height = 34

    for n, r in enumerate(rows, start=1):
        ws.append([n, r['why'], r['root'], r['page'], r['how'], r['ref'],
                   r['closest'], r['matched'], r['quote'], r['context'],
                   r['suggest'], r['suggest_n'],
                   r['fix_ref'], r['fix_text'], r['note']])

    rtl = dict(horizontal='right', vertical='center', wrap_text=True,
               readingOrder=2)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(COLS)):
        for i, cell in enumerate(row):
            kind = COLS[i][2]
            cell.border = border
            if kind in ('ar', 'fill-ar'):
                cell.font = Font(name='Amiri', size=14)
                cell.alignment = Alignment(**rtl)
            elif kind == 'ur':
                cell.font = Font(name='Jameel Noori Nastaleeq', size=12)
                cell.alignment = Alignment(**rtl)
            else:
                cell.alignment = Alignment(vertical='center', wrap_text=True)
            if kind == 'sug':
                cell.fill = fill_sug
            elif kind and kind.startswith('fill'):
                cell.fill = fill_in

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = 'A1:%s%d' % (get_column_letter(len(COLS)), ws.max_row)

    dv = DataValidation(type='custom',
                        formula1='=OR(ISBLANK(M2),ISNUMBER(FIND(":",M2)))',
                        allow_blank=True, showErrorMessage=True)
    dv.errorTitle = 'Use the form S:A'
    dv.error = 'Write the reference as surah:ayah, for example 25:53'
    ws.add_data_validation(dv)
    dv.add('M2:M%d' % ws.max_row)

    restored = st['cited'] + st['repaired'] + st['global']
    info = wb.create_sheet('How to use')
    info.column_dimensions['A'].width = 108
    lines = [
        ('Quran Connect — quotations still to be fixed by hand', True),
        ('', False),
        ('Every quotation in the Fatuhat al-Quran articles that carries a [surah:ayah] reference was', False),
        ('replaced with the authoritative Quran text: %d of %d. The %d rows in "Quotes to fix" are the'
         % (restored, st['refs'], len(rows)), False),
        ('ones that could not be matched safely, so the book\'s own damaged Arabic was left in place.', False),
        ('', False),
        ('What to fill in', True),
        ('  Column M — CORRECT reference (S:A). This is the important one. Write it as 25:53.', False),
        ('  Column N — only if the book quotes something the reference alone will not resolve', False),
        ('              (a partial phrase, two verses run together). Otherwise leave it blank.', False),
        ('  Column O — anything you want me to know about that row.', False),
        ('', False),
        ('Rows where column M is left blank stay exactly as the book printed them.', False),
        ('', False),
        ('What the other columns mean', True),
        ('  Book page — the printed page number in the PDF/Word file, not the PDF viewer\'s page.', False),
        ('  Page found by — "ref on page" means the reference itself was located there; "root article"', False),
        ('              means only the start of that root\'s article could be located.', False),
        ('  Closest ayah we found — our best guess, or the list of ayaat when the wording is ambiguous.', False),
        ('  Matched — how many consonants of the quote were found in that ayah.', False),
        ('  Our suggestion — where the digits of the printed reference, rearranged, point at an ayah', False),
        ('              that really does contain the quote. Green cells are a guess, not a decision:', False),
        ('              check them and copy into column M if you agree.', False),
        ('  Quote as printed — the book\'s own text, still carrying the encoding damage (every fatha is', False),
        ('              written as shadda+fatha), which is why it reads oddly.', False),
        ('', False),
        ('Send the file back once it is filled in and the corrections are folded into the build.', False),
    ]
    for i, (text, bold) in enumerate(lines, start=1):
        c = info.cell(row=i, column=1, value=text)
        c.font = Font(bold=bold, size=12 if i == 1 else 11)
        c.alignment = Alignment(vertical='center')

    wb.save(OUT)


def ingest():
    """Read the filled-in workbook into raw/quote_fixes.json, the file the
    build reads. Existing corrections are kept: a row that has been emptied in
    a later round of the workbook does not delete what was recorded before —
    edit the JSON to do that."""
    from openpyxl import load_workbook

    if not os.path.exists(OUT):
        sys.exit('no workbook to read: ' + OUT)
    fixes = {}
    if os.path.exists(R.FIXES):
        for f in json.load(io.open(R.FIXES, encoding='utf-8'))['fixes']:
            fixes[(f['root'], f['ref'], f['quote'])] = f

    wb = load_workbook(OUT, read_only=True, data_only=True)
    added = changed = 0
    for row in wb['Quotes to fix'].iter_rows(min_row=2, values_only=True):
        row = list(row) + [None] * (15 - len(row))
        root, ref, quote = row[2], row[5], row[8]
        fix_ref = str(row[12] or '').strip()
        fix_text = str(row[13] or '').strip()
        note = str(row[14] or '').strip()
        if not (root and ref) or not (fix_ref or fix_text):
            continue
        key = (str(root).strip(), str(ref).strip(),
               ' '.join(str(quote or '').split()))
        rec = {'root': key[0], 'ref': key[1], 'quote': key[2],
               'fix_ref': fix_ref, 'fix_text': fix_text, 'note': note}
        old = fixes.get(key)
        if old is None:
            added += 1
        elif (old.get('fix_ref'), old.get('fix_text')) != (fix_ref, fix_text):
            changed += 1
        fixes[key] = rec
    wb.close()

    ordered = sorted(fixes.values(), key=lambda f: (f['root'], f['ref']))
    with io.open(R.FIXES, 'w', encoding='utf-8') as f:
        json.dump({'source': os.path.basename(OUT), 'fixes': ordered}, f,
                  ensure_ascii=False, indent=1)
    print('corrections: %d in file (%d new, %d changed)'
          % (len(ordered), added, changed))
    print('output: ' + R.FIXES)
    print('now rebuild:  python scripts/build_root_dictionary.py')


def main():
    if '--ingest' in sys.argv:
        ingest()
        return
    if os.path.exists(OUT) and '--force' not in sys.argv:
        sys.exit('%s already exists. Read it in first with --ingest (any '
                 'corrections in it would otherwise be lost), or pass --force '
                 'to overwrite it.' % OUT)
    rows, st = collect()
    write(rows, st)
    for why, n in collections.Counter(r['why'] for r in rows).most_common():
        print('%-45s %3d' % (why, n))
    print('%-45s %3d' % ('with a suggested reference', sum(1 for r in rows if r['suggest'])))
    print('%-45s %3d / %d' % ('located on a printed page',
                              sum(1 for r in rows if r['page']), len(rows)))
    print('output: %s (%.0f KB)' % (OUT, os.path.getsize(OUT) / 1024))


if __name__ == '__main__':
    main()
